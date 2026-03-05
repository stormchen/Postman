"""
Mantis Bug Tracker 資料撈取與 Excel 匯出工具

功能說明：
- 自動分頁撈取 Mantis API 資料
- 過濾過去 7 天內有更新的 tickets
- 動態攤平 custom_fields 欄位
- 匯出為 Excel 檔案
"""

import requests
import json
from datetime import datetime, timedelta
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
import os
import sys
from typing import List, Dict, Optional
import smtplib
from email.message import EmailMessage
from email.mime.base import MIMEBase
from email import encoders


class MantisExporter:
    """Mantis Bug Tracker 資料撈取與匯出類別"""
    
    def __init__(self):
        """初始化 Mantis Exporter，載入環境變數"""
        # 載入 .env 檔案
        load_dotenv()
        
        # 從環境變數讀取設定
        self.api_base_url = "http://10.210.2.90/mantisbt/api/rest/index.php/issues"
        self.account_id = os.getenv('ACCOUNTID')
        self.token = os.getenv('TOKEN')
        self.filter_id = "734"
        self.page_size = 300
        
        # 從環境變數讀取 SMTP 設定
        self.smtp_server = os.getenv('SMTP_SERVER')
        self.smtp_port = os.getenv('SMTP_PORT')
        self.sender_email = os.getenv('SENDER_EMAIL')
        self.teams_channel_email = os.getenv('TEAMS_CHANNEL_EMAIL')
        
        # 驗證必要的環境變數
        if not self.token:
            print("❌ 錯誤：未找到 TOKEN，請確認 .env 檔案中已設定 TOKEN")
            sys.exit(1)
        # SMTP 基本驗證（若後續發信失敗會在 send_email 顯示）
        if not all([self.smtp_server, self.smtp_port, self.sender_email, self.teams_channel_email]):
            print("⚠ SMTP 設定不完整，Email 功能將無法使用")
        
        print(f"✓ Mantis Exporter 初始化完成")
        print(f"  API 位址：{self.api_base_url}")
        print(f"  Filter ID：{self.filter_id}")
        print(f"  Page Size：{self.page_size}")
        print()
    
    def _get_headers(self) -> Dict[str, str]:
        """
        構建 API 請求的 Header
        
        Returns:
            Dict[str, str]: HTTP Header 字典
        """
        headers = {
            'Authorization': self.token,
            'Content-Type': 'application/json'
        }
        
        # 如果有 ACCOUNTID，也加入 Header
        if self.account_id:
            headers['AccountID'] = self.account_id
        
        return headers
    
    def _parse_datetime(self, datetime_str: str) -> Optional[datetime]:
        """
        解析 ISO 8601 格式的時間字串（含時區資訊）
        
        Args:
            datetime_str: ISO 8601 格式的時間字串，例如 "2026-03-05T12:32:20+08:00"
        
        Returns:
            Optional[datetime]: 解析後的 datetime 物件，失敗時返回 None
        """
        try:
            # 處理 ISO 8601 格式，包含時區資訊
            # Python 3.7+ 可以直接使用 fromisoformat
            if '+' in datetime_str or datetime_str.count('-') > 2:
                # 移除時區資訊，只保留本地時間
                datetime_part = datetime_str.split('+')[0].split('-')[0:3]
                datetime_str_clean = datetime_str[:19]  # 取前 19 個字元 (YYYY-MM-DDTHH:MM:SS)
                return datetime.fromisoformat(datetime_str_clean)
            else:
                return datetime.fromisoformat(datetime_str)
        except Exception as e:
            print(f"⚠ 時間解析失敗：{datetime_str}，錯誤：{e}")
            return None
    
    def _is_within_last_7_days(self, updated_at_str: str) -> bool:
        """
        檢查更新時間是否在過去 7 天內
        
        Args:
            updated_at_str: ISO 8601 格式的時間字串
        
        Returns:
            bool: 是否在過去 7 天內
        """
        updated_at = self._parse_datetime(updated_at_str)
        if not updated_at:
            return False
        
        # 計算 7 天前的時間
        seven_days_ago = datetime.now() - timedelta(days=7)
        
        return updated_at >= seven_days_ago
    
    def _flatten_custom_fields(self, custom_fields: List[Dict]) -> Dict[str, str]:
        """
        攤平 custom_fields 陣列，將其轉換為字典
        
        Args:
            custom_fields: custom_fields 陣列
        
        Returns:
            Dict[str, str]: 攤平後的欄位字典
        """
        flattened = {}
        
        if not custom_fields:
            return flattened
        
        for field_item in custom_fields:
            try:
                # 從 field 物件中提取欄位名稱
                field_name = field_item.get('field', {}).get('name', 'Unknown')
                # 提取欄位值
                field_value = field_item.get('value', '')
                
                # 將欄位名稱與值加入字典
                flattened[field_name] = field_value
            except Exception as e:
                print(f"⚠ Custom field 攤平失敗：{e}")
        
        return flattened
    
    def _extract_issue_data(self, issue: Dict) -> Optional[Dict]:
        """
        從單筆 issue 資料中提取必要的欄位
        
        Args:
            issue: 單筆 issue 資料字典
        
        Returns:
            Optional[Dict]: 提取後的資料字典，若不符合篩選條件則返回 None
        """
        try:
            # 檢查更新時間是否在過去 7 天內
            updated_at_str = issue.get('updated_at', '')
            if not self._is_within_last_7_days(updated_at_str):
                return None
            
            # 提取基本欄位
            extracted_data = {
                'Ticket ID': issue.get('id', ''),
                'Project': issue.get('project', {}).get('name', ''),
                'Status': issue.get('status', {}).get('name', ''),
                'Severity': issue.get('severity', {}).get('name', ''),
                'Category': issue.get('category', {}).get('name', ''),
                'Summary': issue.get('summary', ''),
                'Updated At': updated_at_str[:19] if updated_at_str else '',  # 移除時區資訊
            }
            
            # 新增額外過濾條件（不區分大小寫）
            # 狀態篩選
            allowed_status = {'new', 'assigned'}
            status_val = extracted_data['Status'].lower()
            if status_val not in allowed_status:
                return None
            # 嚴重性篩選
            allowed_severity = {'normal', 'serious', 'critical'}
            severity_val = extracted_data['Severity'].lower()
            if severity_val not in allowed_severity:
                return None
            # 類別篩選
            allowed_category = {'bios', 'bmc', 'general', 'hw'}
            category_val = extracted_data['Category'].lower()
            if category_val not in allowed_category:
                return None
            
            # 攤平 custom_fields
            custom_fields_list = issue.get('custom_fields', [])
            flattened_fields = self._flatten_custom_fields(custom_fields_list)
            
            # 處理 Processing 自訂欄位過濾
            processing_value = flattened_fields.get('Processing', '')
            allowed_processing = {'fae', 'ee', 'bios', 'bmc'}
            if processing_value.lower() not in allowed_processing:
                return None
            
            # 合併基本欄位與 custom_fields
            extracted_data.update(flattened_fields)
            
            return extracted_data
        
        except Exception as e:
            print(f"⚠ Issue 資料提取失敗 (ID: {issue.get('id', 'Unknown')})：{e}")
            return None
    
    def fetch_all_issues(self) -> List[Dict]:
        """
        分頁撈取所有符合條件的 issues
        
        Returns:
            List[Dict]: 符合篩選條件的 issues 資料列表
        """
        all_issues = []
        page = 1
        
        print("🔄 開始撈取 Mantis 資料...")
        
        while True:
            try:
                # 構建 API 請求 URL
                params = {
                    'filter_id': self.filter_id,
                    'page_size': self.page_size,
                    'page': page
                }
                
                # 發送 HTTP GET 請求
                response = requests.get(
                    self.api_base_url,
                    params=params,
                    headers=self._get_headers(),
                    timeout=30
                )
                
                # 檢查 HTTP 狀態碼
                if response.status_code != 200:
                    print(f"❌ API 請求失敗 (Page {page})：狀態碼 {response.status_code}")
                    print(f"   回應內容：{response.text}")
                    break
                
                # 解析 JSON 回應
                response_data = response.json()
                issues = response_data.get('issues', [])
                
                # 若沒有更多資料，結束迴圈
                if not issues or len(issues) == 0:
                    print(f"✓ 已撈取所有資料 (共 {page - 1} 頁)")
                    break
                
                print(f"✓ 已撈取第 {page} 頁 ({len(issues)} 筆資料)")
                
                # 處理本頁的 issues
                for issue in issues:
                    extracted = self._extract_issue_data(issue)
                    if extracted:  # 只保留符合時間篩選條件的資料
                        all_issues.append(extracted)
                
                # 若本頁資料數小於 page_size，表示已到最後一頁
                if len(issues) < self.page_size:
                    print(f"✓ 已撈取所有資料 (共 {page} 頁)")
                    break
                
                page += 1
            
            except requests.exceptions.Timeout:
                print(f"❌ API 請求逾時 (Page {page})")
                break
            except requests.exceptions.RequestException as e:
                print(f"❌ API 請求發生錯誤 (Page {page})：{e}")
                break
            except json.JSONDecodeError:
                print(f"❌ API 回應 JSON 解析失敗 (Page {page})")
                break
            except Exception as e:
                print(f"❌ 未預期的錯誤 (Page {page})：{e}")
                break
        
        print(f"\n📊 撈取完成：共 {len(all_issues)} 筆符合條件的資料\n")
        return all_issues
    
    def export_to_excel(self, issues_data: List[Dict]) -> (bool, Optional[str], int):
        """
        將資料匯出為 Excel 檔案，並設定欄位顯示/隱藏及新增欄位

        Args:
            issues_data: issues 資料列表

        Returns:
            (bool, Optional[str], int):
                是否成功、檔案名稱（若成功）、資料筆數
        """
        # 若沒有資料，不需匯出
        if not issues_data:
            print("⚠ 沒有符合篩選條件的資料，不生成 Excel 檔案")
            return False, None, 0

        try:
            # 轉換為 pandas DataFrame
            df = pd.DataFrame(issues_data)

            # 生成檔案名稱（包含當日日期）
            today = datetime.now().strftime('%Y%m%d')
            filename = f"Mantis_Weekly_Update_{today}.xlsx"

            # 先匯出基本的 Excel 檔案
            df.to_excel(filename, index=False, engine='openpyxl')

            # 使用 openpyxl 進行進階格式設定
            from openpyxl import load_workbook
            from openpyxl.styles import Font

            wb = load_workbook(filename)
            ws = wb.active

            # 定義要保留顯示的欄位
            visible_columns = ['Ticket ID', 'Project', 'Category', 'Summary']

            # 隱藏不需要的欄位
            for col_num, column in enumerate(df.columns, 1):  # 從第1列開始（openpyxl使用1-based indexing）
                if column not in visible_columns:
                    col_letter = ws.cell(row=1, column=col_num).column_letter
                    ws.column_dimensions[col_letter].hidden = True

            # 加入新欄位：BMC Leader 和 Leader 回報確認狀態
            # 找到最後一列的列號
            last_col_num = len(df.columns) + 1

            # 設定新欄位的標題
            ws.cell(row=1, column=last_col_num, value='BMC Leader')
            ws.cell(row=1, column=last_col_num + 1, value='Leader 回報確認狀態')

            # 設定新欄位標題的字體（加粗）
            header_font = Font(bold=True)
            ws.cell(row=1, column=last_col_num).font = header_font
            ws.cell(row=1, column=last_col_num + 1).font = header_font

            # 儲存修改後的檔案
            wb.save(filename)

            print(f"✓ Excel 檔案已成功生成：{filename}")
            print(f"  共 {len(df)} 筆資料，{len(df.columns)} 個欄位")
            print(f"  已隱藏 {len(df.columns) - len(visible_columns)} 個欄位")
            print(f"  已新增 2 個欄位：BMC Leader, Leader 回報確認狀態")

            return True, filename, len(df)

        except Exception as e:
            print(f"❌ Excel 匯出失敗：{e}")
            return False, None, 0

    def send_email(self, attachment_path: str, ticket_count: int) -> None:
        """
        使用 SMTP 發送信件並附加生成的 Excel 檔案
        
        Args:
            attachment_path: Excel 附件的路徑
            ticket_count: 過濾後的 ticket 數量
        """
        # 若 SMTP 設定不完整，跳過
        if not all([self.smtp_server, self.smtp_port, self.sender_email, self.teams_channel_email]):
            print("⚠ SMTP 設定不完整，無法發送 Email")
            return
        
        try:
            msg = EmailMessage()
            msg['Subject'] = "[自動報表] 每週 EIP Ticket 更新狀態(Testing 2026/3/5)"
            msg['From'] = self.sender_email
            msg['To'] = self.teams_channel_email
            msg.set_content(f"本週共有 {ticket_count} 筆 Ticket 更新，詳細請見附件 Excel。")
            
            # 讀取附件並加入信件
            with open(attachment_path, 'rb') as f:
                file_data = f.read()
                file_name = Path(attachment_path).name
                msg.add_attachment(file_data,
                                   maintype='application',
                                   subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                   filename=file_name)
            
            # 建立 SMTP 連線
            server = smtplib.SMTP(self.smtp_server, int(self.smtp_port))
            server.starttls()
            # 若需登入，可在此加入 server.login(user, pass)
            server.send_message(msg)
            server.quit()
            
            print(f"✓ 已將報表發送至 {self.teams_channel_email}")
        except Exception as e:
            print(f"❌ 發送 Email 失敗：{e}")

def main():
    """主程式進入點"""
    print("=" * 60)
    print("Mantis Bug Tracker 資料撈取與 Excel 匯出工具")
    print("=" * 60)
    print()
    
    try:
        # 初始化 Exporter
        exporter = MantisExporter()
        
        # 撈取所有 issues
        issues_data = exporter.fetch_all_issues()
        
        # 匯出為 Excel
        success, filename, count = exporter.export_to_excel(issues_data)
        
        # 若成功生成檔案，嘗試發送 Email
        if success and filename:
            exporter.send_email(filename, count)
        
        print("=" * 60)
        print("✓ 所有流程已完成")
        print("=" * 60)
    
    except Exception as e:
        print(f"❌ 程式執行失敗：{e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
